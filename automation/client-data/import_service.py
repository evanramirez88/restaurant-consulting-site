"""
Client Data Import Service
Watches the import/pending folder and processes files into client folders.
"""

import os
import json
import time
import shutil
import hashlib
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List, Any
from dataclasses import dataclass, asdict
import re

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """Result of an import operation."""
    success: bool
    filename: str
    client_slug: Optional[str]
    destination: Optional[str]
    file_type: str
    error: Optional[str] = None
    extracted_data: Optional[Dict] = None


@dataclass
class ClientMatch:
    """A potential client match for an imported file."""
    slug: str
    confidence: float  # 0-1
    match_reason: str


class FileParser:
    """Base class for file parsers."""
    
    def can_parse(self, filepath: Path) -> bool:
        raise NotImplementedError
    
    def parse(self, filepath: Path) -> Dict[str, Any]:
        raise NotImplementedError
    
    def extract_client_hints(self, filepath: Path, content: Dict) -> List[str]:
        """Extract potential client identifiers from content."""
        return []


class TextParser(FileParser):
    """Parser for .txt and .md files."""
    
    EXTENSIONS = {'.txt', '.md'}
    
    def can_parse(self, filepath: Path) -> bool:
        return filepath.suffix.lower() in self.EXTENSIONS
    
    def parse(self, filepath: Path) -> Dict[str, Any]:
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            return {
                'type': 'text',
                'extension': filepath.suffix,
                'content': content,
                'word_count': len(content.split()),
                'line_count': len(content.splitlines()),
                'preview': content[:500] if len(content) > 500 else content
            }
        except Exception as e:
            return {'type': 'text', 'error': str(e)}
    
    def extract_client_hints(self, filepath: Path, content: Dict) -> List[str]:
        hints = []
        text = content.get('content', '')
        
        # Look for email addresses
        emails = re.findall(r'[\w\.-]+@[\w\.-]+\.\w+', text)
        hints.extend(emails)
        
        # Look for company names (capitalized words)
        # This is a simple heuristic
        companies = re.findall(r'(?:^|\n)([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s*(?:Restaurant|Café|Bistro|Grill|Bar|Pub|Kitchen)', text)
        hints.extend(companies)
        
        return hints


class CSVParser(FileParser):
    """Parser for .csv files."""
    
    def can_parse(self, filepath: Path) -> bool:
        return filepath.suffix.lower() == '.csv'
    
    def parse(self, filepath: Path) -> Dict[str, Any]:
        try:
            import csv
            
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                rows = list(reader)
            
            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []
            
            return {
                'type': 'csv',
                'headers': headers,
                'row_count': len(data_rows),
                'preview': data_rows[:5],
                'columns': len(headers)
            }
        except Exception as e:
            return {'type': 'csv', 'error': str(e)}
    
    def extract_client_hints(self, filepath: Path, content: Dict) -> List[str]:
        hints = []
        headers = content.get('headers', [])
        preview = content.get('preview', [])
        
        # Look for email or company columns
        email_cols = [i for i, h in enumerate(headers) if 'email' in h.lower()]
        company_cols = [i for i, h in enumerate(headers) if any(x in h.lower() for x in ['company', 'business', 'restaurant', 'name'])]
        
        for row in preview:
            for col in email_cols:
                if col < len(row) and row[col]:
                    hints.append(row[col])
            for col in company_cols:
                if col < len(row) and row[col]:
                    hints.append(row[col])
        
        return hints


class JSONParser(FileParser):
    """Parser for .json files."""
    
    def can_parse(self, filepath: Path) -> bool:
        return filepath.suffix.lower() == '.json'
    
    def parse(self, filepath: Path) -> Dict[str, Any]:
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            return {
                'type': 'json',
                'keys': list(data.keys()) if isinstance(data, dict) else None,
                'is_array': isinstance(data, list),
                'item_count': len(data) if isinstance(data, list) else 1,
                'preview': str(data)[:500]
            }
        except Exception as e:
            return {'type': 'json', 'error': str(e)}


class XMLParser(FileParser):
    """Parser for .xml files (SMS exports, etc.)."""
    
    def can_parse(self, filepath: Path) -> bool:
        return filepath.suffix.lower() == '.xml'
    
    def parse(self, filepath: Path) -> Dict[str, Any]:
        try:
            import xml.etree.ElementTree as ET
            
            tree = ET.parse(filepath)
            root = tree.getroot()
            
            # Count elements
            element_counts = {}
            for elem in root.iter():
                tag = elem.tag
                element_counts[tag] = element_counts.get(tag, 0) + 1
            
            return {
                'type': 'xml',
                'root_tag': root.tag,
                'element_counts': element_counts,
                'total_elements': sum(element_counts.values())
            }
        except Exception as e:
            return {'type': 'xml', 'error': str(e)}
    
    def extract_client_hints(self, filepath: Path, content: Dict) -> List[str]:
        hints = []
        
        # For SMS exports, extract from filename
        filename = filepath.stem
        # Pattern: name_sms-YYYYMMDDHHMMSS
        match = re.match(r'^(.+?)_sms-\d+', filename)
        if match:
            hints.append(match.group(1).replace('_', ' '))
        
        return hints


class ExcelParser(FileParser):
    """Parser for .xlsx files."""
    
    def can_parse(self, filepath: Path) -> bool:
        return filepath.suffix.lower() in {'.xlsx', '.xls'}
    
    def parse(self, filepath: Path) -> Dict[str, Any]:
        try:
            # Use openpyxl for xlsx
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheets = wb.sheetnames
            
            sheet_info = {}
            for sheet_name in sheets[:3]:  # First 3 sheets only
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(max_row=6, values_only=True))
                sheet_info[sheet_name] = {
                    'headers': rows[0] if rows else [],
                    'row_count': ws.max_row,
                    'preview': rows[1:6] if len(rows) > 1 else []
                }
            
            wb.close()
            
            return {
                'type': 'excel',
                'sheets': sheets,
                'sheet_count': len(sheets),
                'sheet_info': sheet_info
            }
        except ImportError:
            return {'type': 'excel', 'error': 'openpyxl not installed'}
        except Exception as e:
            return {'type': 'excel', 'error': str(e)}


class PDFParser(FileParser):
    """Parser for .pdf files."""
    
    def can_parse(self, filepath: Path) -> bool:
        return filepath.suffix.lower() == '.pdf'
    
    def parse(self, filepath: Path) -> Dict[str, Any]:
        try:
            # Try PyPDF2 first
            from PyPDF2 import PdfReader
            
            reader = PdfReader(filepath)
            num_pages = len(reader.pages)
            
            # Extract text from first few pages
            text_preview = []
            for i, page in enumerate(reader.pages[:3]):
                text = page.extract_text()
                if text:
                    text_preview.append(text[:500])
            
            return {
                'type': 'pdf',
                'page_count': num_pages,
                'text_preview': '\n---\n'.join(text_preview),
                'has_text': bool(text_preview)
            }
        except ImportError:
            return {'type': 'pdf', 'error': 'PyPDF2 not installed', 'page_count': 0}
        except Exception as e:
            return {'type': 'pdf', 'error': str(e)}


class ImportService:
    """Main import service that watches and processes files."""
    
    def __init__(self, base_path: str, clients_db: Optional[Dict] = None):
        self.base_path = Path(base_path)
        self.pending_path = self.base_path / 'import' / 'pending'
        self.processed_path = self.base_path / 'import' / 'processed'
        self.failed_path = self.base_path / 'import' / 'failed'
        self.clients_path = self.base_path / 'clients'
        
        # Known clients for matching
        self.clients_db = clients_db or {}
        
        # Initialize parsers
        self.parsers: List[FileParser] = [
            TextParser(),
            CSVParser(),
            JSONParser(),
            XMLParser(),
            ExcelParser(),
            PDFParser()
        ]
        
        # Ensure directories exist
        for path in [self.pending_path, self.processed_path, self.failed_path, self.clients_path]:
            path.mkdir(parents=True, exist_ok=True)
    
    def get_parser(self, filepath: Path) -> Optional[FileParser]:
        """Get the appropriate parser for a file."""
        for parser in self.parsers:
            if parser.can_parse(filepath):
                return parser
        return None
    
    def parse_filename_for_client(self, filename: str) -> Optional[ClientMatch]:
        """Try to extract client slug from filename."""
        # Pattern: {client-slug}_{doc-type}_{date}.{ext}
        match = re.match(r'^([a-z0-9-]+)_[a-z_]+_[\d-]+\.\w+$', filename.lower())
        if match:
            slug = match.group(1)
            return ClientMatch(slug=slug, confidence=0.9, match_reason='filename_pattern')
        
        # Try to match against known clients
        for slug in self.clients_db.keys():
            if slug.lower() in filename.lower():
                return ClientMatch(slug=slug, confidence=0.7, match_reason='filename_contains_slug')
        
        return None
    
    def determine_folder_type(self, filepath: Path, parsed_data: Dict) -> str:
        """Determine which subfolder the file belongs in."""
        filename = filepath.name.lower()
        
        # Intel keywords
        if any(kw in filename for kw in ['intel', 'research', 'strategy', 'analysis', 'report']):
            return 'intel'
        
        # Documents keywords
        if any(kw in filename for kw in ['contract', 'agreement', 'invoice', 'receipt', 'legal']):
            return 'documents'
        
        # Communications keywords
        if any(kw in filename for kw in ['sms', 'email', 'message', 'chat', 'call']):
            return 'communications'
        
        # Menu keywords
        if any(kw in filename for kw in ['menu', 'food', 'drink', 'dish']):
            return 'menus'
        
        # Quote keywords
        if any(kw in filename for kw in ['quote', 'estimate', 'proposal', 'pricing']):
            return 'quotes'
        
        # Default based on file type
        file_type = parsed_data.get('type', '')
        if file_type == 'xml':
            return 'communications'  # Likely SMS export
        
        return 'documents'  # Default
    
    def process_file(self, filepath: Path) -> ImportResult:
        """Process a single file."""
        logger.info(f"Processing: {filepath.name}")
        
        # Get parser
        parser = self.get_parser(filepath)
        if not parser:
            return ImportResult(
                success=False,
                filename=filepath.name,
                client_slug=None,
                destination=None,
                file_type='unknown',
                error=f'No parser for file type: {filepath.suffix}'
            )
        
        # Parse file
        try:
            parsed_data = parser.parse(filepath)
        except Exception as e:
            return ImportResult(
                success=False,
                filename=filepath.name,
                client_slug=None,
                destination=None,
                file_type=filepath.suffix,
                error=f'Parse error: {str(e)}'
            )
        
        # Check for parse errors
        if 'error' in parsed_data:
            return ImportResult(
                success=False,
                filename=filepath.name,
                client_slug=None,
                destination=None,
                file_type=parsed_data.get('type', 'unknown'),
                error=parsed_data['error']
            )
        
        # Try to match to a client
        client_match = self.parse_filename_for_client(filepath.name)
        
        # If no match from filename, try content hints
        if not client_match:
            hints = parser.extract_client_hints(filepath, parsed_data)
            for hint in hints:
                # Try to match hint to known clients
                hint_lower = hint.lower().replace(' ', '-')
                for slug in self.clients_db.keys():
                    if hint_lower in slug or slug in hint_lower:
                        client_match = ClientMatch(slug=slug, confidence=0.5, match_reason='content_hint')
                        break
                if client_match:
                    break
        
        # Determine destination folder type
        folder_type = self.determine_folder_type(filepath, parsed_data)
        
        # If we have a client match, move to client folder
        if client_match:
            client_folder = self.clients_path / client_match.slug / folder_type
            client_folder.mkdir(parents=True, exist_ok=True)
            
            dest_path = client_folder / filepath.name
            
            # Handle duplicate filenames
            if dest_path.exists():
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                stem = filepath.stem
                suffix = filepath.suffix
                dest_path = client_folder / f"{stem}_{timestamp}{suffix}"
            
            # Move file
            shutil.move(str(filepath), str(dest_path))
            
            # Create import log
            self._write_import_log(dest_path, client_match, parsed_data)
            
            return ImportResult(
                success=True,
                filename=filepath.name,
                client_slug=client_match.slug,
                destination=str(dest_path),
                file_type=parsed_data.get('type', 'unknown'),
                extracted_data=parsed_data
            )
        else:
            # No client match - move to failed for manual review
            dest_path = self.failed_path / filepath.name
            if dest_path.exists():
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                dest_path = self.failed_path / f"{filepath.stem}_{timestamp}{filepath.suffix}"
            
            shutil.move(str(filepath), str(dest_path))
            
            return ImportResult(
                success=False,
                filename=filepath.name,
                client_slug=None,
                destination=str(dest_path),
                file_type=parsed_data.get('type', 'unknown'),
                error='No client match found - moved to failed for manual review',
                extracted_data=parsed_data
            )
    
    def _write_import_log(self, filepath: Path, client_match: ClientMatch, parsed_data: Dict):
        """Write an import log file alongside the imported file."""
        log_path = filepath.parent / f".{filepath.name}.import_log.json"
        log_data = {
            'imported_at': datetime.now().isoformat(),
            'original_name': filepath.name,
            'client_match': asdict(client_match),
            'parsed_data': parsed_data
        }
        with open(log_path, 'w') as f:
            json.dump(log_data, f, indent=2, default=str)
    
    def scan_pending(self) -> List[ImportResult]:
        """Scan pending folder and process all files."""
        results = []
        
        # Skip README files
        skip_files = {'README.md', 'readme.md', '.gitkeep'}
        
        for filepath in self.pending_path.iterdir():
            if filepath.is_file() and filepath.name not in skip_files:
                result = self.process_file(filepath)
                results.append(result)
                logger.info(f"  Result: {'✓' if result.success else '✗'} {result.filename} -> {result.client_slug or 'NO MATCH'}")
        
        return results
    
    def watch(self, interval: int = 5):
        """Watch the pending folder continuously."""
        logger.info(f"Starting import watch on: {self.pending_path}")
        logger.info("Press Ctrl+C to stop")
        
        try:
            while True:
                results = self.scan_pending()
                if results:
                    logger.info(f"Processed {len(results)} files")
                time.sleep(interval)
        except KeyboardInterrupt:
            logger.info("Import watch stopped")


def main():
    """Main entry point for the import service."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Client Data Import Service')
    parser.add_argument('--base-path', default='.', help='Base path for client-data folder')
    parser.add_argument('--watch', action='store_true', help='Watch mode - continuously monitor')
    parser.add_argument('--interval', type=int, default=5, help='Watch interval in seconds')
    
    args = parser.parse_args()
    
    # Load known clients from JSON if available
    clients_db = {}
    clients_file = Path(args.base_path) / 'clients' / '_clients.json'
    if clients_file.exists():
        with open(clients_file) as f:
            clients_db = json.load(f)
    else:
        # Add known test clients
        clients_db = {
            'crown-anchor': {'name': 'Crown & Anchor', 'company': 'Crown & Anchor'},
        }
    
    service = ImportService(args.base_path, clients_db)
    
    if args.watch:
        service.watch(args.interval)
    else:
        results = service.scan_pending()
        print(f"\nProcessed {len(results)} files:")
        for r in results:
            status = '✓' if r.success else '✗'
            print(f"  {status} {r.filename} -> {r.client_slug or 'NO MATCH'}")
            if r.error:
                print(f"      Error: {r.error}")


if __name__ == '__main__':
    main()
